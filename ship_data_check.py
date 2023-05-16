import pandas as pd
import os
import json
import openpyxl
from datetime import datetime
from datetime import date
from openpyxl.styles import Border , Side , Alignment , PatternFill , Font
from openpyxl.styles.colors import Color

# read data
df = pd.read_excel('./data.xlsx')
wb = openpyxl.Workbook()

# sheet create
ws = wb.create_sheet('Ship ANA')

# no using sheet delete
del wb['Sheet']

# data column write on the first line
ws.append(['Day', 'Time' ,'trailer','Packing Slip Number','Partnumber','qty','hour','Shift'])

# take the datas from original data excel file
for count in range(df.shape[0]):

    # bring JSON as string
    item = df['Json Msg'][count]

    # change string to json ; '{ ... }' -> { ... }
    result = json.loads(item)

    # data find according feature
    if result['detail'][0]['containerno'][0] == '6':
        partnumber = result['detail'][0]['partno']
    else :
        partnumber = result['detail'][0]['detail'][0]['partno']

    trailer = result['trailerno']
    packingslipno = result['packingslipno']

    createdat = result['expecteddatetime']
    date_day = createdat.split('T')[0]
    time = createdat.split('T')[1][:-1]
    time_round = datetime.strptime(date_day+" "+time, '%Y-%m-%d %H:%M:%S')
    
    # item code name -> company secret
    items = []

    if partnumber == items[2] or partnumber == items[3] :
        partnumber = '21 Inch'

    elif partnumber == items[1] :
        partnumber = '20 Inch'
    
    elif partnumber == items[0] :
        partnumber = '19 Inch'
    
    if time_round.hour>=0 and time_round.hour<=6 :
        shift = '3rd Shift'
    elif time_round.hour>=7 and time_round.hour<=15 :
        shift = '1st Shift'
    else :
        shift = '2nd Shift'

    ws.append([date_day,time,trailer,packingslipno,partnumber,1,time_round.hour,shift])


wb.save('./TEST.xlsx')