import os
from openpyxl import load_workbook
import pandas as pd
import datetime
import copy
from openpyxl.utils.cell import get_column_letter
import openpyxl.chart
from openpyxl.chart import BarChart,Reference,Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Font , Border , Side , Alignment , PatternFill
from openpyxl.chart import legend

def edit_maker() :
    fileDir = r"C:\\Users\\suhyu\\Desktop\\공유폴더\\공유\\T사 생산계획\\original\\"
    saveDir = r"C:\\Users\\suhyu\\Desktop\\공유폴더\\공유\\T사 생산계획\\edited\\"
    fileExt = r".xlsx"

    name_list = os.listdir(fileDir)
    saved_list = os.listdir(saveDir)

    item_path = "C:\\Users\\suhyu\\Desktop\\공유폴더\\items.txt"

    years = [2023,2024]

    for xlsx_name in name_list :

        if xlsx_name[:-5]+'_updated'+xlsx_name[-5:] in saved_list:
            pass
        else :
            path = fileDir+xlsx_name
            update_path = saveDir+xlsx_name[:-5]+'_updated'+xlsx_name[-5:]
            wb = load_workbook(path)

            Raw = wb['Raw']

            # item list read from txt file.
            '''
            
            If there is change in the item list, for example added or deleted item ,
            should edit the txt file.

            '''
            items = open(item_path,'r')
            items_read = items.read()
            item_list = items_read.split(',\n')
            items.close()

            for year_take in years:
                    

                # data process
                data_process={}
                m_data_process={}
                for c,(i,j,k) in enumerate(zip(Raw['E'],Raw['G'],Raw['I'])):
                    if c != 0 and datetime.datetime.strptime(j.value,'%m/%d/%Y').year == year_take :
                        
                        # weekly data process
                        cw = datetime.datetime.strptime(j.value,'%m/%d/%Y').isocalendar()[1]
                        if cw not in data_process:
                            data_process[cw] = [0,0,0,0]
                        data_process[cw][item_list.index(i.value)] += k.value
                        
                        # monthly data process <-----------------------------------
                        _month = datetime.datetime.strptime(j.value,'%m/%d/%Y').month
                        if _month not in m_data_process:
                            m_data_process[_month] = [0,0,0,0]
                        m_data_process[_month][item_list.index(i.value)] += k.value

                # weekly data process in a sheet
                wb.create_sheet(str(year_take)+' weekly',1)
                weekly = wb[str(year_take)+' weekly']
                weekly.append(['Calendar week'] + item_list + ['Open Quantity'])

                # monthly data process in a sheet <-----------------------------------
                wb.create_sheet(str(year_take)+' monthly',2)
                monthly = wb[str(year_take)+' monthly']
                monthly.append(['month'] + item_list + ['Open Quantity'])

                # Now there are only 4 items . If new item is going to add , have to add below width line.
                # weekly.column_dimensions["G"].width = WIDTH ...

                WIDTH = 15
                weekly.column_dimensions["A"].width = WIDTH
                weekly.column_dimensions["B"].width = WIDTH
                weekly.column_dimensions["C"].width = WIDTH
                weekly.column_dimensions["D"].width = WIDTH
                weekly.column_dimensions["E"].width = WIDTH
                weekly.column_dimensions["F"].width = WIDTH

                # for montly sheet <-----------------------------------
                monthly.column_dimensions["A"].width = WIDTH
                monthly.column_dimensions["B"].width = WIDTH
                monthly.column_dimensions["C"].width = WIDTH
                monthly.column_dimensions["D"].width = WIDTH
                monthly.column_dimensions["E"].width = WIDTH
                monthly.column_dimensions["F"].width = WIDTH

                t = sorted(list(data_process.keys()))
                for add_cw in t :
                    add_item = [add_cw]
                    add_item += data_process[add_cw]
                    add_item.append(sum(data_process[add_cw]))
                    weekly.append(add_item)

                # for montly sheet <-----------------------------------
                t_m = sorted(list(m_data_process.keys()))
                for add_month in t_m :
                    add_item_m = [add_month]
                    add_item_m += m_data_process[add_month]
                    add_item_m.append(sum(m_data_process[add_month]))
                    monthly.append(add_item_m)

                # border make
                thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

                for row_items in weekly.iter_rows(max_col=weekly.max_column):
                    for item in row_items:
                        item.border = thin_border
                        item.alignment = Alignment(horizontal="center",vertical="center")

                # 2 weeks colored with yellow color
                for color_cells in weekly.iter_rows(min_row=2,max_row=3,min_col=2,max_col=5):
                    for color_cell in color_cells :
                        color_cell.fill = PatternFill(fgColor="FFFF00" , fill_type="solid")

                # for montly sheet <-----------------------------------
                for m_row_items in monthly.iter_rows(max_col=monthly.max_column):
                    for m_item in m_row_items:
                        m_item.border = thin_border
                        m_item.alignment = Alignment(horizontal="center",vertical="center")



                # chart make
                bar_value = Reference(weekly,min_row=1,max_row=5,min_col=2,max_col=5)
                bar_set_categories = Reference(weekly,min_row=2,max_row=5,min_col=1,max_col=1)
                bar_chart = BarChart()
                bar_chart.add_data(bar_value, titles_from_data=True)
                bar_chart.set_categories(bar_set_categories)
                bar_chart.title = "4 weeks forecast"
                bar_chart.type = 'col'
                # chart is the stacked form. For changing should delete the below line.
                bar_chart.grouping = 'stacked'
                bar_chart.overlap = 100
                weekly.add_chart(bar_chart,"G2")


                
                m_bar_value = Reference(monthly, min_row=1, max_row=monthly.max_row, min_col=2, max_col=5)
                m_bar_set_categories = Reference(monthly, min_row=2, max_row=monthly.max_row, min_col=1, max_col=1)
                m_bar_chart = BarChart()
                m_bar_chart.add_data(m_bar_value, titles_from_data=True)
                m_bar_chart.set_categories(m_bar_set_categories)
                m_bar_chart.title = "month forecast"
                m_bar_chart.type = 'col'
                # chart is the stacked form. For changing should delete the below line.
                m_bar_chart.grouping = 'stacked'
                m_bar_chart.overlap = 100

                # m_bar_chart.dataLabels = DataLabelList()
                # m_bar_chart.dataLabels.showVal = True
                # line = m_bar_chart.series[1]
                # line.trendline = Trendline(trendlineType='exp')
                
                m_line_chart = LineChart()
                m_line_value = Reference(monthly, min_row=1, max_row=monthly.max_row, min_col=6, max_col=6)
                m_line_chart.add_data(m_line_value, titles_from_data=True)
                m_line_chart.series[0].graphicalProperties.line.noFill = True
                m_line_chart.dataLabels = DataLabelList(dLblPos='t')
                m_line_chart.dataLabels.showVal = True

                m_line_chart.series[0].trendline = Trendline(trendlineType='exp')
                

                m_bar_chart += m_line_chart
                # m_bar_chart.legend.legendEntry.append(legend.LegendEntry(idx=len(chart.series) - 1, delete=1))
                # m_bar_chart.legend = None
                m_bar_chart.height = 22
                m_bar_chart.width = 44
                monthly.add_chart(m_bar_chart,"G2")

            # save updated excel.
            wb.active = weekly
            wb.save(update_path)
            wb.close()
edit_maker()
