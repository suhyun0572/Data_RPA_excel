import os
from openpyxl import load_workbook
import pandas as pd
import datetime
import copy
from openpyxl.utils.cell import get_column_letter
from openpyxl.chart import BarChart,Reference,Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment , Border , Side
import time
import win32print
import win32com.client

def edit_print():
    # path
    fileDir = r" ... "
    save_path = r" ... "


    name_list = os.listdir(fileDir)
    saved_list = os.listdir(save_path)

    for xlsx_name in name_list :
        if xlsx_name[:-5]+'_updated'+xlsx_name[-5:] in saved_list:
            pass
        else :
            # create a wookbook and sheet
            wb = load_workbook(fileDir+xlsx_name)
            data1 = wb[wb.sheetnames[0]]

            # create a dataframe
            df1 = pd.DataFrame(data1.values)
            df1 = df1.rename(columns=df1.iloc[0]).drop(df1.index[0])
            need_data = df1.loc[:,['Shipment#' , 'Supplier' , 'AppointmentDateTime' , df1.columns[-2] , df1.columns[-1]]]
            need_data = need_data.sort_values(by=['AppointmentDateTime'])

            # save as excel and read once again
            need_data.to_excel(save_path+xlsx_name[:-5]+'_updated'+xlsx_name[-5:],index=False)
            wb = load_workbook(save_path+xlsx_name[:-5]+'_updated'+xlsx_name[-5:])
            ws = wb[wb.sheetnames[0]]

            # set the column width size
            ws.column_dimensions['A'].width = 21.5
            ws.column_dimensions['B'].width = 45.5
            ws.column_dimensions['C'].width = 49
            ws.column_dimensions['D'].width = 18.5 
            ws.column_dimensions['E'].width = 14 
            
            # set the cell styles
            for currentCell_tuple in ws['A1:E'+str(need_data.shape[0]+1)]:
                for currentCell in currentCell_tuple:
                    currentCell.alignment = Alignment(horizontal='center',vertical='center')
                    currentCell.border = Border(left=Side(border_style="thin",color='FF000000'),
                                                right=Side(border_style="thin",color='FF000000'),
                                                top=Side(border_style="thin",color='FF000000'),
                                                bottom=Side(border_style="thin",color='FF000000')
                                                )
            # print area setting
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.print_area = 'A1:E'+str(need_data.shape[0]+1)
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = True
            ws.sheet_properties.pageSetUpPr.fitToPage = True

            # save
            wb.save(save_path+xlsx_name[:-5]+'_updated'+xlsx_name[-5:])

            # if needed, can order to print with the below code
            # os.startfile(save_path+xlsx_name[:-5]+'_updated'+xlsx_name[-5:], "print")

            wb.close()
            
            # time sleep for the printer
            # time.sleep(5)
    



if __name__=="__main__":
    edit_print()